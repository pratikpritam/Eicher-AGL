import logging
import pandas as pd
import yaml
import time
from datetime import datetime, timedelta
from typing import Dict, Any
from google.cloud import bigquery, storage
import smtplib
from email.message import EmailMessage
import tempfile
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)

class EicherHypEmailerAdaptor:
    def __init__(self, config_path: str, secrets: Dict[str, str]):
        with open(config_path, "r") as cp:
            self.config = yaml.safe_load(cp)
        self.objects = self.config["hyp_emailer"]["objects"]
        self.secrets = secrets
        self.project_id = secrets['PROJECT_ID']
        self.bq_client = bigquery.Client(project=self.project_id)
        self.storage = storage.Client()
        self.source_table = secrets["SOURCE_TABLE"]
        self.mapping_table = secrets["MAPPING_TABLE"]
        self.bucket_name = secrets["GCS_BUCKET"]
        self.bucket = self.storage.bucket(self.bucket_name)
        self.smtp_host = self.config["hyp_emailer"]["SMTP_HOST"]
        self.smtp_port = self.config["hyp_emailer"]["SMTP_PORT"]
        self.smtp_user = self.config["hyp_emailer"]["SMTP_USER"]
        self.smtp_password = secrets["SMTP_PASSWORD"]
        self.email_from = self.config["hyp_emailer"]["EMAIL_FROM"]
        self.cc_emails = self.config["hyp_emailer"]["cc_emailer"]

    def _get_target_month(self) -> str:
        today = datetime.today().date()
        if today.day == 1:
            target_date = today.replace(day=1) - timedelta(days=1)
        else:
            target_date = today
        month_year = target_date.strftime('%B %Y')
        # month_year = "June 2026"
        return month_year

    def extract_dataframe(self) -> pd.DataFrame:
        month_year = self._get_target_month()
        query = f"""
        WITH mapping AS (
            SELECT
                CAST(Parent_Dealer_Code AS STRING) AS parent_code,
                Dealer_Name AS mapped_dealer_name,
                CAST(Outlet_Code AS STRING) AS mapped_outlet_id,
                DP_Email_ID,
                CRE_email_id,
                Supervisior_mail_id
            FROM `{self.mapping_table}`
        )
        SELECT
            b.*,
            m.parent_code,
            m.mapped_dealer_name,
            m.mapped_outlet_id,
            m.DP_Email_ID,
            m.CRE_email_id,
            m.Supervisior_mail_id
        FROM `{self.source_table}` b
        LEFT JOIN mapping m
        ON CAST(b.`Outlet ID` AS STRING) = m.mapped_outlet_id
        WHERE b.`Time Period` = '{month_year}'
        """
        logging.info(f"🔹Executing BigQuery to Fetch Data for the Month-Year as: {month_year}")
        df = self.bq_client.query(query).to_dataframe()
        if df.empty:
            logging.warning("⚠️No data fetched.")
            return pd.DataFrame()
        logging.info(f"✅ Data fetched: {len(df)} rows.")
        return df

    def _clean_emails(self, email_str):
        if not email_str:
            return []
        emails = set()
        for e in str(email_str).split(';'):
            e = e.strip().lower()
            if e and '@' in e:
                emails.add(e)
        return list(emails)

    def _prepare_summary_df(self, group):
        summary_df = group.copy()
        summary_df = summary_df[["Outlet ID", "Outlet Name", "Location", "BU", "Total Calls", "Call Answered %","Call Missed %","Call Leads Assigned %","Call Leads Open %","Total Web Leads", "Web Leads Assigned %","Web Leads Open %"]]
        return summary_df

    def _generate_html_table(self, summary_df):
        month_year = self._get_target_month()
        logging.info(f"📅Target month_year selected: {month_year}")
        rows_html = ""
        for _, row in summary_df.iterrows():
            rows_html += f"""
            <tr>
                <td style="text-align:center; vertical-align:middle;">
                    {row['Outlet ID']}
                </td>
                <td style="text-align:center; vertical-align:middle;">
                    {row['Outlet Name']}
                </td>
                <td style="text-align:center; vertical-align:middle;">
                    {row['Location']}
                </td>
                <td style="text-align:center; vertical-align:middle;">
                    {row['BU']}
                </td>
                <td style="text-align:center; vertical-align:middle;">
                    {row['Total Calls']}
                </td>
                <td style="text-align:center; vertical-align:middle;">
                    {row['Call Answered %']}
                </td>
                <td style="text-align:center; vertical-align:middle;">
                    {row['Call Missed %']}
                </td>
                <td style="text-align:center; vertical-align:middle;">
                    {row['Call Leads Assigned %']}
                </td>
                <td style="text-align:center; vertical-align:middle;">
                    {row['Call Leads Open %']}
                </td>
                <td style="text-align:center; vertical-align:middle;">
                    {row['Total Web Leads']}
                </td>
                <td style="text-align:center; vertical-align:middle;">
                    {row['Web Leads Assigned %']}
                </td>
                <td style="text-align:center; vertical-align:middle;">
                    {row['Web Leads Open %']}
                </td>
            </tr>
            """
        html = f"""
        <html>
        <body style="font-family:Calibri; font-size:14px; color:#333333;">
        <p>Hi Team,</p>
    
        <p>Hope you are doing well!</p>
    
        <p>
        Please find below a summary of the Outlet Performance for the period of <b>'{month_year}'-MTD</b>.
        </p>
    
        <table
            border="1"
            cellspacing="0"
            cellpadding="8"
            style="
                border-collapse:collapse;
                font-family:Calibri;
                font-size:12px;
                width:100%;
                border:2px solid #7F8C8D;
                table-layout:auto;
            "
        >
            <!-- MAIN SECTION HEADER -->
            <tr style="font-weight:bold; text-align:center;">
                <!-- BLANK SECTION -->
                <th colspan="4"
                    style="
                        background-color:#DCE6F1;
                        border:2px solid #7F8C8D;
                    ">
                </th>

                <!-- CALL STATUS -->
                <th colspan="5"
                    style="
                        background-color:#B8CCE4;
                        border:2px solid #5B9BD5;
                        font-weight:bold;
                        font-size:14px;
                    ">
                    Call Status
                </th>
    
                <!-- WEB LEAD STATUS -->
                <th colspan="3"
                    style="
                        background-color:#E4DFEC;
                        border:2px solid #C0504D;
                        font-weight:bold;
                        font-size:14px;
                    ">
                    Web Lead Status
                </th>
            </tr>
    
            <!-- COLUMN HEADERS -->
            <tr style="font-weight:bold; text-align:center;">
                <th style="background-color:#DCE6F1;">Outlet ID</th>
                <th style="background-color:#DCE6F1;">Outlet Name</th>
                <th style="background-color:#DCE6F1;">Location</th>
                <th style="background-color:#DCE6F1;">BU</th>
    
                <th style="background-color:#B8CCE4;">Total Calls</th>
                <th style="background-color:#B8CCE4;">Answered %</th>
                <th style="background-color:#B8CCE4;">Missed %</th>
                <th style="background-color:#B8CCE4;">Assigned %</th>
                <th style="background-color:#B8CCE4;">Open %</th>
    
                <th style="background-color:#E4DFEC;">Total Web Lead</th>
                <th style="background-color:#E4DFEC;">Web Leads Assigned %</th>
                <th style="background-color:#E4DFEC;">Web Leads Open %</th>
            </tr>
            <!-- DATA ROWS -->
            {rows_html}
        </table>
    
        <br>
    
        <p>
        Detailed Outlet-wise performance has been attached herewith for your reference.
        </p>
    
        <p>
        Request all the teams to review the performance metrics and align on necessary actionables to improve customer experience.
        </p>
    
        <br>
    
        <p>
        Thanks & Regards,
        <br>
        Data Team
        </p>
    
        <br><br>
    
        <p style="font-style:italic; color:#666666; font-size:14px;">
        -------------|| Note: Please do not reply to this email, as this mailbox is not monitored. ||---------------
        </p>
        </body>
        </html>
        """
        return html

    def _upload_to_gcs(self, file_path, dealer_name):
        today = datetime.today()
        year = today.strftime('%Y')
        month = today.strftime('%m')
        day = today.strftime('%d')
        month_year = self._get_target_month().replace(" ", "_")
        safe_name = (dealer_name.replace(" ", "_").replace(".", ""))
        file_name = (f"{safe_name}_LMS_Adherence_Report_{month_year}_MTD.xlsx")

        gcs_path = (f"hyperlocal_dealer_emailer/{year}/{month}/{day}/{file_name}")
        blob = self.bucket.blob(gcs_path)
        blob.upload_from_filename(file_path)
        logging.info(f"☁️For Audit Purpose, '{file_name}' Uploaded to GCS bucket @: {gcs_path}")
        return f"gs://{self.bucket_name}/{gcs_path}"

    def _send_email(self, emails, dealer_name, html_body, file_path):
        month_year = self._get_target_month()
        msg = EmailMessage()
        msg['Subject'] = (f"Daily hyperlocal performance - {dealer_name}")
        msg['From'] = self.email_from
        msg['To'] = ','.join(emails)
        msg['CC'] = ','.join(self.cc_emails)
        msg.set_content("HTML Email")
        msg.add_alternative(html_body, subtype='html')
        with open(file_path, 'rb') as f:
            msg.add_attachment(
                f.read(),
                maintype='application',
                subtype=(
                    'vnd.openxmlformats-officedocument.'
                    'spreadsheetml.sheet'
                ),
                filename=(f"{dealer_name}_LMS_Adherence_Report_{month_year}_MTD.xlsx")
            )
        all_recipients = emails + self.cc_emails
        with smtplib.SMTP(self.smtp_host, self.smtp_port) as server:
            server.starttls()
            server.login(self.smtp_user, self.smtp_password)
            server.send_message(msg, to_addrs=all_recipients)
        logging.info(f"📧Email Sent to all Receipents for Parent_Dealer: [{dealer_name}]")

    def process_and_send(self, df):
        missing = df[df['mapped_dealer_name'].isna()]
        if not missing.empty:
            logging.info(f"⚠️ Missing mapping rows: {len(missing)}")
        df = df[df['mapped_dealer_name'].notna()]
        grouped = df.groupby(['parent_code', 'mapped_dealer_name'])
        email_sent_count = 0
        for (parent_code, mapped_dealer_name), group in grouped:
            email_sent_count += 1
            if email_sent_count % 10 == 0:
                self._sleep()
            logging.info(f"🔹Running For: [{mapped_dealer_name}]")
            email_set = set()
            for col in ['DP_Email_ID', 'CRE_email_id', 'Supervisior_mail_id']:
                if col in group.columns:
                    for val in group[col].dropna():
                        cleaned = self._clean_emails(val)
                        email_set.update(cleaned)
            emails = list(email_set)
            if not emails:
                logging.info(f"⚠️No Valid Emails: {mapped_dealer_name}")
                continue
            logging.info(f"For The Parent_Dealer [{mapped_dealer_name}], Email Recepient ({len(emails)}) & List as: {emails}")
            summary_df = self._prepare_summary_df(group)
            html_body = self._generate_html_table(summary_df)
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                file_path = tmp.name
            attachment_df = group.drop(
                columns=['DP_Email_ID','CRE_email_id','Supervisior_mail_id','parent_code','mapped_dealer_name','mapped_outlet_id'],
                errors='ignore'
            )
            attachment_df = attachment_df[[
                'Time Period', 'Outlet ID', 'Outlet Name', 'Location', 'State', 'Region', 'BU',
                'Total Calls', 'Answered', 'IVR Drop', 'Missed', 'Offline', 'Call Missed %',
                'Duplicate Call Leads', 'Follow Up Scheduled Call Leads', 'In Discussion Call leads', 'Call Junk Leads', 'Call Open Leads', 'Call Leads Assigned', 'Call Leads Open %',
                'Total Web Leads', 'Duplicate Web Leads', 'Follow Up Scheduled Web Leads', 'In Discussion Web Leads', 'Web Junk Leads', 'Web Open Leads', 'Web Leads Assigned', 'Web Leads Open %'
            ]]

            attachment_df.columns = [
                'Time Period', 'Outlet ID', 'Outlet Name', 'Location', 'State', 'Region', 'BU',
                'Total Calls', 'Answered', 'IVR_Drop', 'Missed', 'Offline', 'Missed %',
                'Duplicate', 'Follow Up Scheduled', 'In Discussion', 'Junk', 'Open', 'Assigned', 'Open %',
                'Total Web Lead', 'Duplicate', 'Follow Up Scheduled', 'In Discussion', 'Junk', 'Open', 'Assigned', 'Open %'
            ]
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                attachment_df.to_excel(
                    writer,
                    sheet_name='Detailed Report',
                    index=False,
                    startrow=1
                )
                workbook = writer.book
                worksheet = writer.sheets['Detailed Report']
                generic_fill = PatternFill(start_color="EAF2F8", end_color="EAF2F8", fill_type="solid")
                call_fill = PatternFill(start_color="EBF5EB", end_color="EBF5EB", fill_type="solid")

                call_lead_fill = PatternFill(start_color="FCF3CF",end_color="FCF3CF",fill_type="solid")
                web_fill = PatternFill(start_color="E4DFEC",end_color="E4DFEC",fill_type="solid")
                yellow_fill = PatternFill(start_color="FFFF00",end_color="FFFF00",fill_type="solid")

                thick_side = Side(border_style="thick",color="000000")
                thin_side = Side(border_style="thin",color="000000")
                thick_border = Border(left=thick_side,right=thick_side,top=thick_side,bottom=thick_side)
                thin_border = Border(left=thin_side,right=thin_side,top=thin_side,bottom=thin_side)

                worksheet.merge_cells('A1:G1')
                # worksheet['H1'] = ""
                
                worksheet.merge_cells('H1:M1')
                worksheet['H1'] = "Call Status"

                worksheet.merge_cells('N1:T1')
                worksheet['N1'] = "Call Lead Status"

                worksheet.merge_cells('U1:AB1')
                worksheet['U1'] = "Web Lead Status"

                for cell in worksheet[1]:
                    cell.font = Font(bold=True,size=12,color="000000")
                    cell.alignment = Alignment(horizontal='center',vertical='center')
                    cell.border = thick_border
                for col in range(8, 14):
                    worksheet.cell(1, col).fill = call_fill
                for col in range(14, 21):
                    worksheet.cell(1, col).fill = call_lead_fill
                for col in range(21, 29):
                    worksheet.cell(1, col).fill = web_fill
                for cell in worksheet[2]:
                    cell.font = Font(bold=True,size=10)
                    cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                    cell.border = thin_border

                for col in range(1, 8):
                    worksheet.cell(2, col).fill = generic_fill
                for col in range(8, 14):
                    worksheet.cell(2, col).fill = call_fill
                for col in range(14, 21):
                    worksheet.cell(2, col).fill = call_lead_fill
                    worksheet.cell(2, col).font = Font(bold=True,color="000000")
                for col in range(21, 29):
                    worksheet.cell(2, col).fill = web_fill
                for row in worksheet.iter_rows(min_row=3,max_row=worksheet.max_row):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center',vertical='center')
                        cell.border = thin_border

                yellow_cols = [13,20,28]
                for col in yellow_cols:
                    for row in range(3,worksheet.max_row + 1):
                        worksheet.cell(row=row, column=col).fill = yellow_fill
                separator_cols = [7, 13, 20, 28]
                for sep_col in separator_cols:
                    for row in range(1, worksheet.max_row + 1):
                        worksheet.cell(row=row, column=sep_col).border = Border(right=thick_side)

                widths = {
                    'A': 14, 'B': 10, 'C': 42, 'D': 16, 'E': 14, 'F': 12, 'G': 8,
                    'H': 12, 'I': 10, 'J': 10, 'K': 10, 'L': 10, 'M': 10,
                    'N': 14, 'O': 20, 'P': 14, 'Q': 10, 'R': 10, 'S': 10, 'T': 10,
                    'U': 16, 'V': 10, 'W': 20, 'X': 14, 'Y': 10, 'Z': 10, 'AA': 10, 'AB': 10
                }
                for col, width in widths.items():
                    worksheet.column_dimensions[col].width = width
                worksheet.freeze_panes = "A3"

            self._upload_to_gcs(file_path, mapped_dealer_name)
            try:
                self._send_email(
                    emails=emails,
                    dealer_name=mapped_dealer_name,
                    html_body=html_body,
                    file_path=file_path
                )
            except Exception as e:
                logging.warning(f"❌Failed For: {mapped_dealer_name}: {e}")
            os.remove(file_path)

    def _sleep(self, seconds: int = 30):
        logging.info(f"😴Sleeping {seconds}s for Email Rate Limits")
        time.sleep(seconds)

    def _get_obj_cfg(self, object_name: str) -> Dict[str, Any]:
        obj = next((o[object_name] for o in self.objects if object_name in o), None)
        if not obj or not obj.get("active", False):
            raise ValueError(f"{object_name} not active")
        return obj

    def get_dataframe(self, object_name: str):
        logging.info(f"Starting Extraction: {object_name}")
        self._get_obj_cfg(object_name)
        df = self.extract_dataframe()
        if df.empty:
            return pd.DataFrame
        logging.info("🔹Starting Email Processing after Extracting Dataframa.")
        self.process_and_send(df)
        logging.info(f"✅Completed Email Scheduler for: {object_name}")
        return df